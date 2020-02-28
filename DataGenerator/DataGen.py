import openpyxl

def data_generator():
    wk =openpyxl.open("C:\\Users\\Dell Laptop\\EndtoEndAutomation\\DataGenerator\\Tdata1.xlsx")

    sh = wk.active

    r = sh.max_row
    c = sh.max_column
    li =[]
    for i in range(1, r+1):
        lil = []
        for j in range(1, c+1):
            un = sh.cell(i, j)
            lil.append(un.value)
        li.append(lil)
    return li

